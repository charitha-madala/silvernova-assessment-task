import os
import logging
import PyPDF2
import docx
import extract_msg
import openpyxl  # Using openpyxl instead of pandas for Excel
from pathlib import Path

logger = logging.getLogger('markdown-extractor')

class MarkdownExtractor:
    """
    A utility class for extracting content from various document types and 
    converting them to markdown format for RAG embeddings.
    """    
    def __init__(self, input_dir="documents", output_dir="markdown"):
        self.input_dir = input_dir
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def process_directory(self):
        """Process all supported documents in the input directory."""
        processed_count = 0
        skipped_count = 0
        
        for file_path in Path(self.input_dir).glob('**/*'):
            if file_path.is_file():
                extension = file_path.suffix.lower()
                if extension in ['.pdf', '.docx', '.xlsx', '.msg', '.txt', '.md']:
                    self.process_document(file_path)
                    processed_count += 1
                else:
                    logger.debug(f"Skipping unsupported file: {file_path}")
                    skipped_count += 1
        
        logger.info(f"Processing complete. Processed {processed_count} files, skipped {skipped_count} files.")
    
    def process_document(self, file_path):
        """Process a single document based on its file type."""
        try:
            file_path = Path(file_path)
            extension = file_path.suffix.lower()
            
            logger.info(f"Processing {file_path}")
            
            # Extract content based on file type
            if extension == '.pdf':
                content = self.extract_pdf(file_path)
            elif extension == '.docx':
                content = self.extract_docx(file_path)
            elif extension == '.xlsx':
                content = self.extract_excel(file_path)
            elif extension == '.msg':
                content = self.extract_email(file_path)
            elif extension in ['.txt', '.md']:
                content = self.extract_text(file_path)
            else:
                logger.warning(f"Unsupported file type: {file_path}")
                return         
            
            self.save_as_markdown(file_path, content)
            
        except Exception as e:
            logger.error(f"Error processing {file_path}: {str(e)}")
    
    def extract_pdf(self, file_path):
        """Extract content from PDF files."""
        sections = []        
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file) 
                for i, page in enumerate(pdf_reader.pages):
                    text = page.extract_text()
                    if text.strip(): 
                        sections.append({
                            'title': f"Page {i+1}",
                            'content': text.strip()
                        })
        except Exception as e:
            logger.error(f"Error extracting PDF content from {file_path}: {str(e)}")
        
        return sections
    
    def extract_docx(self, file_path):
      """Extract content from .docx files."""
      text = ""
      try:
          doc = docx.Document(file_path)
          for para in doc.paragraphs:
               if para.text:
                text += para.text + "\n"
          return [{'title': 'Document Content', 'content': text}]
      except Exception as e:
          logger.error(f"Error extracting text from {file_path}: {e}")
      
  
    def extract_excel(self, file_path):
        """Extract content from Excel files using openpyxl instead of pandas."""
        sections = []        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]                
                markdown_lines = []                
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row == 0 or max_col == 0:
                    markdown_lines.append("*Empty sheet*")
                else:
                    header_row = "|"
                    separator_row = "|"                    
                    for col in range(1, max_col + 1):
                        cell_value = sheet.cell(row=1, column=col).value
                        header_row += f" {cell_value if cell_value is not None else ''} |"
                        separator_row += " --- |"                    
                    markdown_lines.append(header_row)
                    markdown_lines.append(separator_row)                    
          
                    for row in range(2, max_row + 1):
                        data_row = "|"
                        row_values = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]                        
                        # Skip rows where all values are None
                        if all(cell is None for cell in row_values):
                            continue
                        
                        for cell_value in row_values:
                            data_row += f" {cell_value if cell_value is not None else ''} |"
                        
                        markdown_lines.append(data_row)                
                sections.append({
                    'title': f"Sheet: {sheet_name}",
                    'content': '\n'.join(markdown_lines)
                })
                
        except Exception as e:
            logger.error(f"Error extracting Excel content from {file_path}: {str(e)}")            
            sections.append({
                'title': 'Excel Content',
                'content': f"Could not process Excel file: {str(e)}"
            })        
        return sections
       
    def extract_email(self, file_path):
      """Extract content from .msg email files and create markdown sections for each email."""
      try:
          msg = extract_msg.Message(file_path)

          subject = msg.subject
          sender = msg.sender
          receiver = msg.to
          date = msg.date
          body = msg.body
          
          body = "\n".join([line.strip() for line in body.split("\n") if line.strip()])
          
          sections = [{
              'title': f"Subject: {subject}",
              'content': f"From: {sender}\nTo: {receiver}\nDate: {date}\n\n{body}"
          }]
          
          # Return the list of sections
          return sections

      except Exception as e:
          print(f"Error extracting email content from {file_path}: {str(e)}")
          return {}

    def extract_text(self, file_path):
      """Extract content from text files (.txt, .md)."""
      try:
          with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
              content = file.read()
          
          return [{
            'title': 'Text Content',
            'content': content
        }]

      except Exception as e:
          print(f"Error extracting text content from {file_path}: {str(e)}")
          return ""
    
    def save_as_markdown(self, file_path, sections):
        """Save extracted content as markdown."""
        base_file_name = file_path.stem
        md_path = Path(self.output_dir) / f"{base_file_name}.md"
        
        metadata = {
            "source_file": file_path.name,
            "file_type": file_path.suffix.lower(),
        }
        markdown_lines = [
            f"# Document: {file_path.name}",
            "",
            "## Metadata",
            f"- **Source**: {metadata['source_file']}",
            f"- **Type**: {metadata['file_type']}",
            ""
        ]
        for section in sections:
            markdown_lines.extend([
                f"## {section['title']}",
                "",
                section['content'],
                ""
            ])
        with open(md_path, "w", encoding="utf-8", errors="replace") as f:
            f.write("\n".join(markdown_lines))

        
    